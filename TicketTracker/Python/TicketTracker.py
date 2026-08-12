"""
ticket_tracker.py

Automates the Jira First Response Time report — single workbook version.

WORKFLOW
--------
1. Export tickets from Jira as CSV.
2. Paste the raw export into the "raw_input" sheet of Ticket_Tracker.xlsx
   (headers: Issue key, Summary, Created, Reporter, Assignee, Status, Resolved).
3. Run:
     python ticket_tracker.py Ticket_Tracker.xlsx
4. The script:
   - Reads raw_input
   - Appends new tickets to the "Incidents" sheet in the management format
   - Skips any ticket already present in Incidents (safe to re-run)
   - Leaves "Acknowledged Time (IST)" blank for new rows (filled in manually)
   - Deletes the raw_input sheet entirely once done

COLUMN MAPPING (raw_input -> Incidents)
-----------------------------------------
Issue key   -> Ticket No                (plain text)
Issue key   -> Jira                     (=HYPERLINK formula to the Jira ticket)
Summary     -> Ticket Details
Created     -> Created Time (IST)       (copied as-is, Jira now exports in IST)
--          -> Acknowledged Time (IST)  (left blank, filled in manually)
Resolved    -> Resolved Time (IST)      (copied as-is)
Reporter    -> Reporter
Assignee    -> Assigned To
Status      -> Status
"""

import sys
from pathlib import Path
from copy import copy

import pandas as pd
from openpyxl import load_workbook

# ---- Config -----------------------------------------------------------
RAW_SHEET = "raw_input"
TARGET_SHEET = "Incidents"
JIRA_BASE_URL = "https://lectragroup.atlassian.net/browse/"

REQUIRED_RAW_COLUMNS = [
    "Issue key", "Summary", "Created", "Reporter", "Assignee", "Status", "Resolved",
]

TEMPLATE_COLUMNS = [
    "Ticket No", "Ticket Details", "Jira", "Created Time (IST)",
    "Acknowledged Time (IST)", "Resolved Time (IST)",
    "Reporter", "Assigned To", "Status",
]

DATE_COLUMNS = {"Created Time (IST)", "Resolved Time (IST)"}
DATE_NUMBER_FORMAT = "dd-mm-yyyy hh:mm"


def load_raw(tracker_path: Path) -> pd.DataFrame:
    df = pd.read_excel(tracker_path, sheet_name=RAW_SHEET)
    missing = [c for c in REQUIRED_RAW_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"'{RAW_SHEET}' sheet is missing required column(s): {missing}"
        )
    return df


def build_report_rows(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    out["Ticket No"] = df["Issue key"]
    out["Ticket Details"] = df["Summary"]
    out["Jira"] = df["Issue key"].apply(
        lambda key: f'=HYPERLINK("{JIRA_BASE_URL}{key}","{key}")'
    )
    out["Created Time (IST)"] = pd.to_datetime(df["Created"], dayfirst=True)
    out["Acknowledged Time (IST)"] = None  # filled in manually after running
    out["Resolved Time (IST)"] = pd.to_datetime(df["Resolved"], dayfirst=True)
    out["Reporter"] = df["Reporter"]
    out["Assigned To"] = df["Assignee"]
    out["Status"] = df["Status"]
    return out[TEMPLATE_COLUMNS]


def append_to_incidents(report_df: pd.DataFrame, wb) -> tuple[int, int]:
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"'{TARGET_SHEET}' sheet not found in the workbook")
    ws = wb[TARGET_SHEET]

    header = [cell.value for cell in ws[1]]
    while header and header[-1] is None:
        header.pop()
    if header != TEMPLATE_COLUMNS:
        raise ValueError(
            f"Header row in '{TARGET_SHEET}' does not match the expected "
            f"template columns.\nExpected: {TEMPLATE_COLUMNS}\nFound:    {header}"
        )

    existing_tickets = {
        ws.cell(row=r, column=1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    }

    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last_row = r
    style_row = last_row if last_row >= 2 else 2

    added, skipped = 0, 0
    next_row = last_row + 1

    for _, row in report_df.iterrows():
        if row["Ticket No"] in existing_tickets:
            skipped += 1
            continue

        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            src_cell = ws.cell(row=style_row, column=col_idx)
            value = row[col_name]
            if pd.isna(value):
                value = None
            dst_cell = ws.cell(row=next_row, column=col_idx, value=value)
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.alignment = copy(src_cell.alignment)
            if col_name in DATE_COLUMNS:
                dst_cell.number_format = DATE_NUMBER_FORMAT
            else:
                dst_cell.number_format = src_cell.number_format

        next_row += 1
        added += 1

    return added, skipped


def delete_raw_sheet(wb):
    if RAW_SHEET in wb.sheetnames:
        del wb[RAW_SHEET]


def main():
    if len(sys.argv) != 2:
        print("Usage: python ticket_tracker.py <Ticket_Tracker.xlsx>")
        sys.exit(1)

    tracker_path = Path(sys.argv[1])

    df = load_raw(tracker_path)
    report_df = build_report_rows(df)

    wb = load_workbook(tracker_path)
    added, skipped = append_to_incidents(report_df, wb)
    delete_raw_sheet(wb)
    wb.save(tracker_path)

    print(f"Done. Added {added} new ticket(s), skipped {skipped} duplicate(s).")
    print(f"'{RAW_SHEET}' sheet removed. Report saved to: {tracker_path}")


if __name__ == "__main__":
    main()
